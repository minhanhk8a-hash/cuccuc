import httpx
import random
import requests
from telegram import Update, InputFile
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import nest_asyncio
from google.cloud import vision
import io
import os
import pytz
from tenacity import retry, stop_after_attempt, wait_fixed
import asyncio
import re
import gdown
from dotenv import load_dotenv
import json
import pandas as pd
from telegram.ext import Application, CommandHandler
from openpyxl import load_workbook
import warnings
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Border, Side
from googletrans import Translator
import zipfile
import xlrd
from datetime import datetime, timedelta
import unicodedata

# Tải tệp từ Google Drive
env_file_id = "1vEA7XWP7-Q5y7JVT4Z5CcC0kr5fflcX0"
json_file_id = "1VZ4RiQH4oqS4NJAAvB5Ec-bY2vOLh70h"

env_file_path = "/content/.env"
json_file_path = "/content/config.json"
api_limits_file_path = "/content/api_limits.json"
api_usage_file_path = "/content/api_usage.json"

# Tải các file
gdown.download(f"https://drive.google.com/uc?id={env_file_id}", env_file_path, quiet=False)
gdown.download(f"https://drive.google.com/uc?id={json_file_id}", json_file_path, quiet=False)

# Hardcode API_LIMITS
API_LIMITS = {
    "-1003292792359": 100,  # abc
    "-100245": 1000,   # test đá
    "-1002375088024": 1000,   # go
    "-1003126414098": 500,  # vip
    "-1002510856954": 100
}
DEFAULT_API_LIMIT = 100

# Lưu API_LIMITS vào api_limits.json (local)
with open(api_limits_file_path, "w", encoding="utf-8") as f:
    json.dump(API_LIMITS, f, indent=4)

# Kiểm tra và tạo api_usage.json nếu không tồn tại
if not os.path.exists(api_usage_file_path):
    with open(api_usage_file_path, "w") as f:
        json.dump({}, f)

# Kiểm tra config.json
if os.path.exists(json_file_path):
    print(f"Tệp JSON tải thành công: {json_file_path}")
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = json_file_path
    try:
        vision_client = vision.ImageAnnotatorClient()
        print("Vision API Client OK.")
    except Exception as e:
        print(f"Lỗi Vision API Client: {e}")
else:
    print(f"Lỗi: {json_file_path} không tồn tại")
    raise FileNotFoundError(f"{json_file_path} không tồn tại")

# Nạp .env
load_dotenv(env_file_path)

# Biến môi trường
port = os.getenv("PORT")
db_host = os.getenv("DB_HOST")
print(f"Port: {port}, DB Host: {db_host}")

# Kiểm tra key
GROK_API_KEY = os.getenv("GROK_API_KEY")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
if not GROK_API_KEY or not TELEGRAM_TOKEN:
    raise ValueError("GROK_API_KEY/TELEGRAM_TOKEN lỗi")
else:
    print("GROK_API_KEY và TELEGRAM_TOKEN OK.")

# Đọc config.json
with open(json_file_path, "r") as json_file:
    config = json.load(json_file)
api_key = config.get("apiKey")
print(f"API Key: {api_key}")

# Hàm đọc/ghi api_usage.json
api_lock = asyncio.Lock()

async def load_api_usage():
    async with api_lock:
        if os.path.exists(api_usage_file_path):
            with open(api_usage_file_path, "r") as f:
                return json.load(f)
        return {}

async def save_api_usage(usage):
    async with api_lock:
        with open(api_usage_file_path, "w") as f:
            json.dump(usage, f, indent=4)

async def check_api_limit(chat_id):
    usage = await load_api_usage()
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    today = datetime.now(tz).strftime("%Y-%m-%d")
    chat_id = str(chat_id)
    limit = API_LIMITS.get(chat_id, DEFAULT_API_LIMIT)

    if chat_id not in usage or usage[chat_id].get("date") != today:
        usage[chat_id] = {"date": today, "count": 0}

    if usage[chat_id]["count"] >= limit:
        return False, f"Xong phim, Quota ảnh cháy túi rồi bro, tự xử nha, rảnh thì chém gió với tao hoặc thử skill khác nha 😎"

    usage[chat_id]["count"] += 1
    await save_api_usage(usage)
    return True, None

warnings.filterwarnings("ignore", category=UserWarning, module="pdfminer")
warnings.filterwarnings("ignore", category=UserWarning, message="Conditional Formatting extension is not supported and will be removed")

# Danh sách chat ID được phép
GROUP_TRANSLATION_ONLY = -1002468518838
ALLOWED_CHAT_IDS = [
    "-1003292792359",  # abc
    "-1002403996935",  # test đá
    "-1002375088024",  # go
    "-1003126414098",  # vip
    "-1002510856954",
]

ALLOWED_PRIVATE_USER_IDS = [
    "7037013131",
    "6936833955",
    "7392677521",
    # "1122334455",
]

# Câu trả lời ngẫu nhiên khi chỉ tag bot
random_replies = [
    "Nói lẹ đi, tao còn bận chơi!",
    "Gì đấy? Mày lại hỏi cái gì nữa?",
    "Mày tag cái gì? Nói lẹ lên!",
    "Cái gì? Đang ngủ",
    "Gì mày?",
]

PRIVATE_BLOCK_REPLIES = [
    "Mày có xinh gái không mà đòi nói chuyện với tao?! 😎",
    "Ơ mày, mày tưởng tao là con bot công cộng à? Đẳng cấp chưa tới thì đứng ngoài nha mày! 💥",
    "Mày nhắn tao làm gì? Muốn tán tao à? Xin lỗi, tao đã có chủ rồi! 😏",
    "Ê mày, tao có người yêu rồi, đi ra xếp hàng xếp nha mày! 😜",
]

used_private_replies = []

# Biến lưu trạng thái tìm kiếm
pending_searches = {}  # {user_key: {"action": "search", "files": {file_name: file_path}, "search_text": text}}

# Lịch sử trò chuyện
conversation_history = {}  # Cho chat cá nhân
group_conversation_history = {}  # Cho nhóm
group_conversation_timestamps = {}

# Hàm xử lý khi chỉ tag bot
def handle_tag_only():
    global recent_replies
    available_replies = [r for r in random_replies if r not in recent_replies]
    if not available_replies:
        available_replies = random_replies
    reply = random.choice(available_replies)
    recent_replies.append(reply)
    if len(recent_replies) > 5:
        recent_replies.pop(0)
    return reply

# Hàm xóa lịch sử nhóm sau 1 giờ
def cleanup_group_history():
    current_time = datetime.now()
    for chat_id in list(group_conversation_history.keys()):
        if current_time - group_conversation_timestamps.get(chat_id, current_time) > timedelta(hours=1):
            del group_conversation_history[chat_id]
            del group_conversation_timestamps[chat_id]
            print(f"Đã xóa lịch sử nhóm cho chat_id: {chat_id}")

# Hàm phân tích văn bản từ ảnh
def detect_text_from_image(file_path):
    try:
        with io.open(file_path, 'rb') as image_file:
            content = image_file.read()
        image = vision.Image(content=content)
        response = vision_client.text_detection(image=image)
        texts = response.text_annotations
        if not texts:
            return "Có cái chữ nào trong ảnh đâu, mày đùa tao à."

        detected_text = texts[0].description.strip()
        lines = detected_text.split("\n")

        # Giữ nguyên lines cho pattern và tên
        full_text = "\n".join(lines)

        # Xử lý số riêng để tìm cặp số (giữ nguyên logic cũ)
        processed_lines_numbers = []
        for line in lines:
            def process_numbers(match):
                sequence = match.group()
                return re.sub(r'[\D\.]', '', sequence)
            processed_line = re.sub(r'(\d[\d\s\-\>\.]*\d)', process_numbers, line)
            processed_lines_numbers.append(processed_line)
        full_text_numbers = "\n".join(processed_lines_numbers)
        all_numbers = re.findall(r'\b\d{11,17}\b', full_text_numbers)

        def find_matching_pairs(numbers):
            valid_lengths = {(15, 14), (14, 15), (12, 15), (15, 12), (13, 14), (14, 13),
                             (13, 11), (11, 13), (14, 14), (12, 16), (16, 12), (14, 17), (17, 14)}
            pairs = []
            for i in range(len(numbers)):
                for j in range(i + 1, len(numbers)):
                    if (len(numbers[i]), len(numbers[j])) in valid_lengths:
                        pairs.append((numbers[i], numbers[j]))
            return pairs
        matching_pairs = find_matching_pairs(all_numbers)

        patterns = {
            "pattern_90_91_92_93": r"\b(?:10|11|12|99)\d{10}\b",
            "pattern_tax": r"\bTAX\w*\b",
            "pattern_EVN": r"\b(EVN\w{8})\b",
            "pattern_VN8P": r"\b(VN8P\w{8})\b",
            "pattern_NYY25": r"\b(NYY25\w{8})\b",
            "pattern_QQ886": r"\b(QQ886\w{8})\b",
            "pattern_XP101": r"\b(XP101\w{8})\b",
            "pattern_VN24": r"\b(VN24\w{8})\b",
            "pattern_OPG25": r"\b(OPG25\w{8})\b",
            "pattern_X666": r"\b(X666\w{8})\b",
            "pattern_U8PAY": r"\b(U8PAY\w{8})\b",
            "pattern_D666": r"\b(D666\w{8})\b",
            "pattern_vp": r"\b([vV][pP]\w{6})\b",
            "pattern_g2p": r"\b[gG][2][pP]\w*\b",
            "pattern_g8b": r"\b[gG][8][bB]\w*\b",
            "pattern_popmart_numbers": r"\bpopmart\b.*?\b\d{7}\b",
            "pattern_vip": r"\b[vV][iI][pP]\w{6}\b",
            "pattern_01MM": r"\b01MM\w*\b",
            "pattern_XJ": r"\b[xX][jJ]\w{6}\b",
            "pattern_TP": r"\b[tT][pP]\w{6}\b",
            "pattern_ft09": r"\bFT09\d{8}\b",
            "pattern_ft10": r"\bFT08\d{8}\b",
            "pattern_ft24": r"\bFT24\d{6}\b",
            "pattern_YLR": r"\bYLR\s[a-zA-Z0-9]{10}\b",
            "pattern_P09": r"\bP09\d{21}\b",
            "pattern_24W": r"\b24W[a-zA-Z0-9]{8}\b",
            "pattern_SP25": r"\bSP25[a-zA-Z0-9]{12}\b",
            "pattern_P10": r"\bP08\d{21}\b",
            "zing_serial": r"\b(?=.*[A-Z])(?=.*\d)[A-Z0-9]{12}\b",
            "zing_code": r"\b(?=.*[A-Z])(?=.*\d)[A-Z0-9]{9}\b"
        }

        # Tìm cặp seri và mã thẻ Zing trước
        zing_pairs = []
        zing_serials = re.findall(patterns["zing_serial"], full_text)
        zing_codes = re.findall(patterns["zing_code"], full_text)
        if zing_serials and zing_codes:
            for serial in zing_serials:
                for code in zing_codes:
                    # Kiểm tra vị trí gần nhau
                    serial_lines = [i for i, line in enumerate(lines) if serial in line]
                    code_lines = [i for i, line in enumerate(lines) if code in line]
                    if any(abs(sl - cl) <= 1 for sl in serial_lines for cl in code_lines):
                        zing_pairs.append((serial, code))

        # Danh sách các pattern cần tự động viết hoa toàn bộ khi match
        uppercase_patterns = [
            "pattern_EVN",
            "pattern_NYY25",
            "pattern_QQ886",
            "pattern_XP101",
            "pattern_VN24",
            "pattern_OPG25",
            "pattern_X666",
            "pattern_U8PAY",
            "pattern_VN8P",
            "pattern_D666"
        ]

        # Tìm các matches
        matches = []
        for key, pattern in patterns.items():
            if key not in ["zing_serial", "zing_code"]:
                if key in uppercase_patterns:
                    # Tìm không phân biệt hoa/thường và trả về dạng in hoa hoàn toàn
                    found = re.findall(pattern, full_text, flags=re.IGNORECASE)
                    matches += [code.upper() for code in found]
                elif key == "pattern_vp":
                    vp_found = re.findall(pattern, full_text, flags=re.IGNORECASE)
                    matches += vp_found
                else:
                    matches += re.findall(pattern, full_text)

        # Chỉ tìm kiếm tên nếu phát hiện pattern_90_91_92_93
        momo_name = None
        random_message = None
        if any(re.match(r"\b(?:10|11|12|99)\d{10}\b", match) for match in matches):
            skip_names = {"Tên Ví MoMo", "Tên gợi nhớ", "Số điện thoại", "Đặt tên gợi nhớ ở", "LỜI NHẮN",
                          "Chuyển thêm", "Danh mục", "Giải trí", "Danh mục", "Danh mục", "Chưa phân loại", "Ăn uống", "Chưa phân loại v", "Chợ, siêu thị", "Mua sắm", "Đặt tên gợi nhớ ơ", "Xem biên nhận giao dịch"}

            def is_valid_name(text):
                return (any(c.isalpha() for c in text) and
                        not text.isdigit() and
                        len([c for c in text if c.isalpha()]) >= 2 and
                        not re.match(r'\b[A-Z]{8,9}\b', text))

            skip_pattern = r'.*\b(' + '|'.join(re.escape(name) for name in skip_names) + r')\b.*'
            skip_regex = re.compile(skip_pattern, re.IGNORECASE)

            # === ƯU TIÊN 1: TÌM TÊN SAU "MOMO-TKTH" (SỬA CHỖ NÀY) ===
            momo_tkth_pattern = re.compile(r'\bMOMO-TKTH\s+([^\n\r]+?)(?=\s*Số thẻ|Tin nhắn|\n|$)', re.IGNORECASE)
            momo_tkth_match = momo_tkth_pattern.search(full_text)

            # DEBUG: In ra để kiểm tra
            print("DEBUG MOMO-TKTH MATCH:", momo_tkth_match.group(0) if momo_tkth_match else "KHÔNG TÌM THẤY")

            if momo_tkth_match:
                potential_name = momo_tkth_match.group(1).strip()
                print("DEBUG TÊN TÌM ĐƯỢC:", potential_name)
                if is_valid_name(potential_name) and not skip_regex.search(potential_name):
                    momo_name = potential_name
                    print("→ DÙNG TÊN:", momo_name)

            # === LUÔN CHẠY PHẦN DƯỚI NẾU CHƯA CÓ TÊN ===
            if momo_name is None:
                start_idx = -1
                end_idx = -1
                for i, line in enumerate(lines):
                    if "Miễn phí" in line:
                        start_idx = i + 1
                    if "*******" in line:
                        end_idx = i

                if start_idx >= 0 and end_idx > start_idx:
                    for i in range(start_idx, end_idx):
                        potential_name = lines[i].strip()
                        if is_valid_name(potential_name) and not skip_regex.match(potential_name):
                            momo_name = potential_name
                            break

                if momo_name is None and end_idx > 0:
                    potential_name = lines[end_idx - 1].strip()
                    if is_valid_name(potential_name) and not skip_regex.match(potential_name):
                        momo_name = potential_name

            # === TÌM 6 CHỮ CÁI ===
            message_candidates = re.findall(r'\b[a-zA-Z]{6}\b', full_text)
            if message_candidates:
                for candidate in reversed(message_candidates):
                    if candidate.islower():
                        random_message = candidate
                        break
                    elif (candidate.startswith("I") and
                          all(c.isalpha() for c in candidate) and
                          not any(c.isupper() for c in candidate[1:])):
                        random_message = "l" + candidate[1:]
                        break

        result = []
        # Thêm cặp seri/mã thẻ Zing
        for serial, code in zing_pairs:
            result.append(f"Seri: {serial} - Mã thẻ: {code}")
        # Thêm các cặp số khác
        if matching_pairs:
            result += [f"{pair[0]} - {pair[1]}" for pair in matching_pairs]
        if matches:
            result += matches
        if momo_name:
            result.append(f"{momo_name}")
        if random_message:
            result.append(f"{random_message}")
        if result:
            return "\n".join(result)
        else:
            # Trả về toàn bộ văn bản như mã ban đầu
            return "\n".join(lines)

    except Exception as e:
        return f"Lỗi rồi, gửi lại đi mày. {e}"

# Hàm xử lý phản hồi từ Grok 3
async def get_grok_response(prompt, chat_id, user_id, is_group_context=False):
    try:
        if is_group_context:
            if chat_id not in group_conversation_history:
                group_conversation_history[chat_id] = [
                    {
                        "role": "system",
                        "content": "Mày là trợ lý siêu lầy, trả lời ngắn, bựa, xưng 'mày' và 'tao'. Hiểu ngữ cảnh nhóm, trả lời dựa trên cuộc trò chuyện chung."
                    }
                ]
            history = group_conversation_history[chat_id]
            group_conversation_timestamps[chat_id] = datetime.now()
        else:
            user_key = f"{chat_id}_{user_id}"
            if user_key not in conversation_history:
                conversation_history[user_key] = [
                    {
                        "role": "system",
                        "content": "Mày là trợ lý siêu lầy, trả lời chỉ 1-2 câu, tối đa 30 từ, bựa, xưng 'mày' và 'tao'. Không lan man, không nghiêm túc! Nếu hỏi dữ liệu phức tạp, trả lời đầy đủ"
                    }
                ]
            history = conversation_history[user_key]

        history.append({"role": "user", "content": prompt})
        if len(history) > 5:
            history = history[-5:]

        headers = {"Authorization": f"Bearer {GROK_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "grok-3",
            "messages": history,
            "temperature": 0.7,
            "max_tokens": 500
        }
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.x.ai/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            response.raise_for_status()
            bot_response = response.json()["choices"][0]["message"]["content"].strip()

        if is_group_context:
            group_conversation_history[chat_id] = history + [{"role": "assistant", "content": bot_response}]
            group_conversation_timestamps[chat_id] = datetime.now()
        else:
            conversation_history[user_key] = history + [{"role": "assistant", "content": bot_response}]

        return bot_response
    except Exception as e:
        return f"Lỗi khi gọi API Grok 3: {e}"

# Hàm xử lý file Excel
def process_excel_with_format(file_path):
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        def extract_patterns(text):
            if pd.isna(text):
                return None
            # Loại bỏ khoảng trắng và xuống dòng trước khi xử lý regex
            cleaned_text = str(text).replace(' ', '').replace('\n', '')
            patterns = [
                r'QAU[Z]\w{11}',# Matches 1IQ, 1IR, 1IS, 1IT, 1IU, 1IV, 1IX, 1IY + 11 chars
                r'1J[ABCDEF]\w{11}',
                r'CTLNHIDI\d{15}',
                r'P09\d{21}',
                r'P10\d{21}',
                r'W199\d{16}',
                r'W200\d{16}',
                r'TAXI\w{6}',
                r'NYY\w{6}',
                r'G2PAY\w{4}',
                r'G8B\w{4}',
                r'XP101\w{4}',
                r'FT24\d{6}',
                r'FT10\d{8}',
                r'FT09\d{8}'
            ]
            # Tìm tất cả các chuỗi khớp với patterns trong cleaned_text
            matches = []
            for pattern in patterns:
                matches.extend(re.findall(pattern, cleaned_text))

            return ', '.join(matches) if matches else ''

        df['Extracted_Data'] = df.apply(
            lambda row: next(
                (extract_patterns(row[col]) for col in df.columns if extract_patterns(row[col])), None
            ),
            axis=1
        )
        workbook = load_workbook(file_path)
        sheet = workbook.active
        sheet.cell(row=1, column=len(df.columns) + 1, value="Extracted_Data")
        for i, row in df.iterrows():
            sheet.cell(row=i + 2, column=len(df.columns) + 1, value=row['Extracted_Data'])
        workbook.save(file_path)
        print(f"File Excel đã được xử lý: {file_path}")
        return file_path
    except FileNotFoundError:
        print("Lỗi: Không tìm thấy file.")
    except Exception as e:
        print(f"Lỗi khi xử lý file Excel: {e}")
    return None

# Hàm chuyển đổi PDF sang Excel
def pdf_to_excel(pdf_path, excel_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = []
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    all_text.extend(table)
            if all_text:
                df = pd.DataFrame(all_text[1:], columns=all_text[0])

                def clean_text(text):
                    if pd.isna(text):
                        return text
                    text_single_line = str(text).replace('\n', ' ').replace('\r', ' ')
                    text_no_spaces = re.sub(r'(?<=\d) (?=\d)', '', text_single_line)
                    return text_no_spaces.strip()

                for col in df.columns:
                    df[col] = df[col].apply(clean_text)
                df.to_excel(excel_path, index=False, sheet_name='Sheet1')
                wb = load_workbook(excel_path)
                ws = wb['Sheet1']
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = thin_border
                wb.save(excel_path)
                print(f"Chuyển đổi PDF sang Excel thành công: {excel_path}")
                processed_file = process_excel_with_format(excel_path)
                if processed_file:
                    print(f"File Excel đã được xử lý: {processed_file}")
                else:
                    print("Không thể xử lý file Excel sau khi chuyển đổi.")
                return excel_path
            else:
                print("Không tìm thấy bảng trong file PDF.")
                return None
    except Exception as e:
        print(f"Lỗi khi chuyển đổi PDF sang Excel: {e}")
        return None

# Hàm định dạng số
def format_number(number):
    try:
        cleaned = str(number).replace(',', '').replace('+', '').replace('-', '').strip()
        if '.' in cleaned:
            cleaned = cleaned.split('.')[0]
        num = int(cleaned)
        return f"{num:,}"  # Trả về số với dấu phẩy
    except (ValueError, TypeError) as e:
        print(f"Lỗi định dạng số: {e}, giá trị: {number}")
        return str(number)

# Hàm kiểm tra file Excel hợp lệ
def is_valid_xlsx(file_path):
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            return True
    except zipfile.BadZipFile:
        return False

# Hàm chuyển đổi .xls sang .xlsx
def convert_xls_to_xlsx(xls_path, xlsx_path):
    try:
        df = pd.read_excel(xls_path, engine='xlrd')
        df.to_excel(xlsx_path, index=False, engine='openpyxl')
        print(f"Đã chuyển đổi {xls_path} sang {xlsx_path}")
        return xlsx_path
    except Exception as e:
        print(f"Lỗi khi chuyển đổi .xls sang .xlsx: {e}")
        return None

# Hàm tìm kiếm trong Excel
def search_in_excel(file_path, search_text):
    try:
        if not is_valid_xlsx(file_path):
            if file_path.endswith('.xls'):
                xlsx_path = file_path.replace('.xls', '_converted.xlsx')
                converted_path = convert_xls_to_xlsx(file_path, xlsx_path)
                if converted_path:
                    file_path = converted_path
                else:
                    raise ValueError("File Excel không hợp lệ hoặc không thể chuyển đổi!")
            else:
                raise ValueError("File Excel không phải định dạng .xlsx hợp lệ!")

        df = pd.read_excel(file_path, engine='openpyxl')
        matches = []
        count = 0

        for row_idx in range(len(df)):
            for col in df.columns:
                cell_value = str(df.at[row_idx, col])
                if search_text.lower() in cell_value.lower():
                    count += 1
                    amount = None
                    for col in df.columns:
                        try:
                            value = df.at[row_idx, col]
                            if pd.notna(value):
                                cleaned_value = str(value).replace('+', '').strip()
                                if '.' in cleaned_value:
                                    cleaned_value = cleaned_value.split('.')[0]
                                cleaned_value = cleaned_value.replace(',', '')
                                if cleaned_value.isdigit():
                                    num_value = int(cleaned_value)
                                    if num_value % 1000 == 0 and num_value > 0:
                                        amount = format_number(num_value)
                                        break
                        except:
                            continue
                    matches.append((row_idx + 2, amount))
        found = count > 0
        return found, count, matches
    except Exception as e:
        print(f"Lỗi khi tìm kiếm trong file Excel: {e}")
        return False, 0, []
    finally:
        if 'converted_path' in locals() and converted_path and os.path.exists(converted_path):
            try:
                os.remove(converted_path)
                print(f"Đã xóa file chuyển đổi tạm: {converted_path}")
            except Exception as e:
                print(f"Lỗi khi xóa file chuyển đổi tạm: {e}")

# Hàm xử lý danh sách số (sửa để trả về chỉ số, định dạng + số)
def process_numbers(text):
    try:
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        even_numbers = []
        odd_numbers = []
        other_numbers = []

        for line in lines:
            # Loại bỏ khoảng trắng, dấu phẩy và phần .00
            cleaned_line = line.replace(',', '').replace(' ', '').replace('.00', '')
            number_str = cleaned_line.lstrip('+-')
            if not number_str.replace(',', '').isdigit():
                continue  # Bỏ qua nếu không phải số hợp lệ
            number = int(number_str.replace(',', ''))
            formatted_number = format_number(number)  # Sử dụng format_number để thêm dấu phẩy

            # Phân loại số
            if number < 3_300_000 or number > 350_000_000:
                other_numbers.append(formatted_number)
            elif number % 1_000_000 == 0:
                even_numbers.append(formatted_number)
            elif number % 500_000 == 0 and number % 1_000_000 != 0:
                odd_numbers.append(formatted_number)
            else:
                other_numbers.append(formatted_number)

        # Tạo các tin nhắn (không có tiêu đề)
        messages = []
        if even_numbers:
            messages.append("\n".join(even_numbers))
        if odd_numbers:
            messages.append("\n".join(odd_numbers))
        if other_numbers:
            messages.append("\n".join(other_numbers))

        return messages if messages else ["Không tìm thấy số hợp lệ!"]
    except Exception as e:
        return [f"Lỗi xử lý số: {e}"]

# Lệnh /search
async def search_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    user_id = str(update.message.from_user.id)
    user_key = f"{chat_id}_{user_id}"

    if chat_id not in ALLOWED_CHAT_IDS:
        await update.message.reply_text("Nhóm này không được phép đâu mày!")
        return

    if user_key in pending_searches:
        old_files = pending_searches[user_key].get("files", {})
        for file_path in old_files.values():
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    print(f"Đã xóa file tìm kiếm cũ: {file_path}")
                except Exception as e:
                    print(f"Lỗi khi xóa file tìm kiếm cũ: {e}")

    pending_searches[user_key] = {"action": "search", "files": {}, "search_text": None}
    await update.message.reply_text("Quăng tao cái file! Nhớ kèm tên nha mày.")

# Lệnh /e
async def e_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    user_id = str(update.message.from_user.id)
    user_key = f"{chat_id}_{user_id}"

    if chat_id not in ALLOWED_CHAT_IDS:
        await update.message.reply_text("Nhóm này không được phép đâu mày!")
        return
    if user_key not in pending_searches or not pending_searches[user_key].get("files"):
        await update.message.reply_text("Chưa có file để tìm kiếm. Gửi /search và file trước")
        return

    args = context.args
    if not args:
        if user_key in pending_searches:
            old_files = pending_searches[user_key].get("files", {})
            for file_path in old_files.values():
                if file_path and os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        print(f"Đã xóa file tìm kiếm: {file_path}")
                    except Exception as e:
                        print(f"Lỗi khi xóa file tìm kiếm: {e}")
            del pending_searches[user_key]
        await update.message.reply_text("Đã hủy tìm kiếm. Gửi /search để chọn file mới!")
        return

    if len(args) < 2:
        await update.message.reply_text("Sai cú pháp! Dùng: /e <tên_file> <nội_dung>")
        return
    file_name = args[0]
    search_text = " ".join(args[1:]).strip()

    files = pending_searches[user_key]["files"]
    if file_name not in files:
        await update.message.reply_text(f"File '{file_name}' không tồn tại! Các file hiện có: {', '.join(files.keys())}")
        return

    pending_searches[user_key]["search_text"] = search_text
    file_path = files[file_name]
    found, count, matches = search_in_excel(file_path, search_text)
    if found:
        if count == 1:
            row, amount = matches[0]
            amount_text = f": {amount}" if amount else "Không tìm thấy số tiền hợp lệ"
            reply = f"'{search_text}' trong '{file_name}': Nhận được! {amount_text}"
        else:
            reply = f"'{search_text}' trong '{file_name}': Nhận được! ({count} lần)\n"
            for row, amount in matches:
                amount_text = f": {amount}" if amount else "Không tìm thấy số tiền hợp lệ"
                reply += f"- Hàng {row}: {amount_text}\n"
    else:
        reply = f"'{search_text}' trong '{file_name}': Chưa nhận!"
    await update.message.reply_text(reply)

# Lệnh /checklimit
async def check_limit_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    if chat_id not in ALLOWED_CHAT_IDS:
        await update.message.reply_text("Nhóm này không được phép đâu mày!")
        return

    try:
        with open(api_limits_file_path, "r", encoding="utf-8") as f:
            API_LIMITS = json.load(f)
        if os.path.exists(api_usage_file_path):
            with open(api_usage_file_path, "r") as f:
                api_usage = json.load(f)
        else:
            api_usage = {}

        tz = pytz.timezone('Asia/Ho_Chi_Minh')
        today = datetime.now(tz).strftime("%Y-%m-%d")
        limit = API_LIMITS.get(chat_id, DEFAULT_API_LIMIT)
        usage = api_usage.get(chat_id, {})
        used_count = usage.get("count", 0) if usage.get("date") == today else 0
        remaining = limit - used_count

        reply = f"Giới hạn của nhóm: {limit}\nĐã dùng: {used_count}\nCòn lại: {remaining}"
        await update.message.reply_text(reply)
    except FileNotFoundError:
        await update.message.reply_text("Lỗi: Không tìm thấy file giới hạn API!")
    except Exception as e:
        await update.message.reply_text(f"Lỗi khi kiểm tra giới hạn: {e}")

# Hàm xử lý dữ liệu Excel dạng text
def process_excel_text(text):
    try:
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        records = []
        errors = []

        for idx, line in enumerate(lines, 1):
            try:
                parts = re.split(r'\s+', line.strip())
                if len(parts) < 6:
                    errors.append(f"Dòng {idx}: Không đủ cột, yêu cầu ít nhất 6 cột (mã P, mã gd, thông tin bank)")
                    continue

                amount = parts[-1]
                try:
                    int(amount.replace(',', ''))  # Kiểm tra amount là số
                except ValueError:
                    errors.append(f"Dòng {idx}: Cột cuối không phải số hợp lệ: {amount}")
                    continue

                col1, col2, col3, col4 = parts[:4]
                name_parts = parts[4:-1]
                name = ' '.join(name_parts)
                records.append([col1, col2, col3, col4, name, amount])
            except Exception as e:
                errors.append(f"Dòng {idx}: Lỗi xử lý: {str(e)}")
                continue

        messages = []
        for idx, record in enumerate(records, 1):
            message = (
                f"{record[0]}\n"
                f"{record[1]}\n\n"
                f"{record[2]}\n"
                f"{record[3]}\n"
                f"{record[4]}\n"
                f"{record[5]}\n\n"
                f"{idx}"
            )
            messages.append(message)

        return messages, errors
    except Exception as e:
        return None, [f"Lỗi tổng quát khi xử lý dữ liệu: {str(e)}"]

# Lệnh /dh
async def dh_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    if chat_id not in ALLOWED_CHAT_IDS:
        await update.message.reply_text("Nhóm này không được phép đâu mày!")
        return

    user_message = update.message.text
    if not user_message:
        await update.message.reply_text("Paste dữ liệu copy từ Excel sau lệnh /dh nha mày!")
        return

    # Tách lệnh /dh ra khỏi dòng đầu tiên
    lines = user_message.split('\n')
    data_lines = []

    # Xử lý dòng đầu tiên: loại bỏ /dh và lấy phần còn lại
    first_line = lines[0].strip()
    if first_line.startswith('/dh'):
        # Lấy phần sau /dh (nếu có)
        data = first_line[len('/dh'):].strip()
        if data:
            data_lines.append(data)

    # Thêm các dòng còn lại
    data_lines.extend([line.strip() for line in lines[1:] if line.strip()])

    if not data_lines:
        await update.message.reply_text("Dữ liệu trống hoặc không đúng, gửi lại nha mày!")
        return

    data_text = '\n'.join(data_lines)
    messages, errors = process_excel_text(data_text)

    if errors:
        error_message = "Có lỗi trong dữ liệu:\n" + "\n".join(errors[:5])  # Giới hạn hiển thị 5 lỗi
        await update.message.reply_text(error_message)

    if messages:
        if len(messages) > 20:
            await update.message.reply_text("Dữ liệu quá lớn! Chỉ xử lý tối đa 20 đơn.")
            messages = messages[:20]

        for message in messages:
            await update.message.reply_text(message)
            await asyncio.sleep(0.5)
    else:
        await update.message.reply_text("Lỗi rồi, không có dữ liệu hợp lệ để xử lý!")

# Danh sách từ khóa
FUNCTION_KEYWORDS = [
    "làm được gì", "có thể làm gì", "làm gì được", "mày làm được"
]

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    user_id = str(update.message.from_user.id)

    # Kiểm tra quyền: private → user_id, group → chat_id
    if update.message.chat.type == "private":
        if user_id not in ALLOWED_PRIVATE_USER_IDS:
            await update.message.reply_text("Mày không được phép dùng lệnh này đâu!")
            return
    else:
        if chat_id not in ALLOWED_CHAT_IDS:
            await update.message.reply_text("Nhóm này không được phép đâu mày!")
            return

    # Danh sách chức năng của bot
    help_text = (
        "Tao là Cục Đá, một trợ lý đa năng, tích hợp AI xin sò, và cả tá skill bá cháy! 😜 \n\n"
        "🔥 Chém gió: Ra xã hội làm ăn bươn chải, liều thì ăn nhiều, không liều thì ăn ít, muốn thành công phải chấp nhận đắng cay ngọt bùi, còn muốn cười rụng hàm thì kiếm tao 😎\n\n"
        "📸 Xử lý ảnh: Quăng tao cái ảnh, tao trích xuất seri, mã thẻ, tên, nội dung, mã giao dịch, muốn gì có nấy, đủ cả! 😋 \n\n"
        "📑 Xử lý file: Gửi file Excel hoặc PDF, tao hô biến, tách dữ liệu, moi hết những gì mày cần, chuyển PDF sang Excel 📁 \n\n"
        "🔍 Tìm kiếm trong file: Gõ /search, chọn file, rồi /e <tên_file> <nội_dung> để tìm dữ liệu 🔎 \n\n"
        "⏰ Tin nhắn tự động: Gào rú đúng giờ, không trượt phát nào, chuẩn như cơm mẹ nấu 🐷 \n\n"
        "📋 Xử lý dữ liệu: Gõ /dh copy dữ liệu gửi tao, tao format lại đẹp hơn crush mày chỉnh ảnh! 😍\n\n"
        "📊 Kiểm tra giới hạn: Dùng /checklimit để test độ may mắn 🐔 \n\n"
        "🔒 Bảo mật: Tao chỉ chơi với nhóm được duyệt, ngoài list tao next, bảo mật căng đét luôn mày! 😎 \n\n"
        "Tag tao, hoặc trả lời tin nhắn của tao để chém gió. Có gì cứ hỏi, tao cân hết! 💪"
    )

    await update.message.reply_text(help_text)

# Hàm để thêm lệnh /help
def add_help_handler(application):
    application.add_handler(CommandHandler("help", help_command))

# Hàm xử lý tin nhắn
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.message.chat_id)
    user_id = str(update.message.from_user.id)
    bot_username = '@' + context.bot.username
    user_message = update.message.text

    # === KHAI BÁO TRƯỚC KHI DÙNG ===
    is_private = update.message.chat.type == "private"
    is_group_chat = update.message.chat.type in ["group", "supergroup"]

    # === XỬ LÝ NHÓM DỊCH THUẬT ===
    if chat_id == str(GROUP_TRANSLATION_ONLY):
        if user_message:
            translator = Translator()
            detected = await translator.detect(user_message)
            if detected.lang == "km":
                translated = await translator.translate(user_message, dest="vi")
                translated_text = translated.text if translated else "Không thể dịch được."
                await context.bot.send_message(chat_id=chat_id, text=f"Dịch từ KHMER:\n{translated_text}")
        return

    # === XỬ LÝ SỐ TRONG PRIVATE CHAT (chỉ user được phép) ===
    if is_private:
        if user_message and re.search(r'^\s*[+-]?\d+(,\d{3})*(\.\d{2})?\s*$', user_message, re.MULTILINE):
            messages = process_numbers(user_message)
            for message in messages:
                await context.bot.send_message(chat_id=chat_id, text=message)
                await asyncio.sleep(0.5)
            return

    # === KIỂM TRA QUYỀN TRUY CẬP ===
    is_private = update.message.chat.type == "private"
    is_group_chat = update.message.chat.type in ["group", "supergroup"]

    if is_private:
        if user_id not in ALLOWED_PRIVATE_USER_IDS:
            global used_private_replies  # ← BẮT BUỘC PHẢI CÓ DÒNG NÀY

            # Nếu đã dùng hết → reset lại
            if len(used_private_replies) >= len(PRIVATE_BLOCK_REPLIES):
                used_private_replies = []

            # Lấy danh sách chưa dùng
            available_replies = [r for r in PRIVATE_BLOCK_REPLIES if r not in used_private_replies]

            # Chọn ngẫu nhiên 1 câu chưa dùng
            reply = random.choice(available_replies)
            used_private_replies.append(reply)

            await update.message.reply_text(reply)
            return
    else:
        # Nhóm: chỉ xử lý nếu nằm trong ALLOWED_CHAT_IDS
        if chat_id not in ALLOWED_CHAT_IDS:
            return  # Bỏ qua nhóm không được phép

    # Xử lý file
    if update.message.document:
        document = update.message.document
        file_name = document.file_name
        if file_name.endswith('.pdf'):
            file_path = f"temp_{file_name}"
            try:
                new_file = await context.bot.get_file(document.file_id)
                await new_file.download_to_drive(file_path)
                excel_file_path = file_path.replace('.pdf', '.xlsx')
                converted_file = pdf_to_excel(file_path, excel_file_path)
                if converted_file:
                    with open(converted_file, "rb") as f:
                        await context.bot.send_document(chat_id=chat_id, document=InputFile(f, filename=excel_file_path))
                else:
                    await context.bot.send_message(chat_id=chat_id, text="Không thể chuyển đổi file PDF thành Excel.")
            except Exception as e:
                print(f"Lỗi khi xử lý file PDF: {e}")
                await context.bot.send_message(chat_id=chat_id, text=f"Không thể xử lý file PDF, gửi lại nha mày: {e}")
            finally:
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        print(f"Đã xóa file tạm: {file_path}")
                    except Exception as e:
                        print(f"Lỗi khi xóa file tạm: {e}")
                if os.path.exists(excel_file_path):
                    try:
                        os.remove(excel_file_path)
                        print(f"Đã xóa file kết quả: {excel_file_path}")
                    except Exception as e:
                        print(f"Lỗi khi xóa file kết quả: {e}")
            return
        elif file_name.endswith(('.xlsx', '.xls')):
            file_path = f"temp_{file_name}"
            try:
                new_file = await context.bot.get_file(document.file_id)
                await new_file.download_to_drive(file_path)

                if file_name.endswith('.xlsx') and not is_valid_xlsx(file_path):
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text="File Excel không hợp lệ hoặc bị hỏng! Vui lòng gửi file .xlsx hoặc .xls khác."
                    )
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            print(f"Đã xóa file tạm: {file_path}")
                        except Exception as e:
                            print(f"Lỗi khi xóa file tạm: {e}")
                    return

                user_key = f"{chat_id}_{user_id}"
                if user_key in pending_searches and pending_searches[user_key]["action"] == "search":
                    caption = update.message.caption
                    if not caption or not caption.strip():
                        await context.bot.send_message(
                            chat_id=chat_id,
                            text="Phải gửi kèm tên file trong caption (ví dụ: 'a1')!"
                        )
                        if os.path.exists(file_path):
                            try:
                                os.remove(file_path)
                                print(f"Đã xóa file tạm: {file_path}")
                            except Exception as e:
                                print(f"Lỗi khi xóa file tạm: {e}")
                        return
                    file_base_name = caption.strip()
                    pending_searches[user_key]["files"][file_base_name] = file_path
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text=f"Đã nhận file '{file_base_name}'. Dùng /e <tên_file> <nội_dung> để tìm kiếm."
                    )
                    return

                output_file = process_excel_with_format(file_path)
                if output_file:
                    with open(output_file, "rb") as f:
                        await context.bot.send_document(chat_id=chat_id, document=InputFile(f, filename=output_file))
                else:
                    await context.bot.send_message(chat_id=chat_id, text="Không thể xử lý file Excel của bạn. File có thể hỏng hoặc không đúng định dạng.")
            except Exception as e:
                print(f"Lỗi khi xử lý file Excel: {e}")
                await context.bot.send_message(chat_id=chat_id, text=f"Không thể xử lý file Excel: {e}")
            finally:
                if user_key not in pending_searches or file_path not in pending_searches[user_key].get("files", {}).values():
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            print(f"Đã xóa file tạm: {file_path}")
                        except Exception as e:
                            print(f"Lỗi khi xóa file tạm: {e}")
                if os.path.exists("filtered_data.xlsx"):
                    try:
                        os.remove("filtered_data.xlsx")
                        print("Đã xóa file kết quả: filtered_data.xlsx")
                    except Exception as e:
                        print(f"Lỗi khi xóa file kết quả: {e}")
            return
        else:
            print("File không phải là PDF hoặc Excel.")
            await context.bot.send_message(chat_id=chat_id, text="Vui lòng gửi file PDF hoặc Excel.")
            return

    # Xử lý ảnh
    if update.message.photo:
        allowed, error_msg = await check_api_limit(chat_id)
        if not allowed:
            await context.bot.send_message(chat_id=chat_id, text=error_msg)
            return

        file_path = "temp_image.jpg"
        try:
            photo_file = await update.message.photo[-1].get_file()
            await photo_file.download_to_drive(file_path)
            try:
                detected_text = detect_text_from_image(file_path)
                await context.bot.send_message(chat_id=chat_id, text=f"Văn bản trong ảnh: \n\n{detected_text}")
            except Exception as e:
                await context.bot.send_message(chat_id=chat_id, text=f"Không thể phân tích ảnh: {e}")
        finally:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Lỗi khi xóa tệp ảnh: {e}")
        return

    # Xử lý tin nhắn văn bản
    if user_message is None:
        return

    # Kiểm tra câu hỏi về chức năng
    user_message_lower = user_message.lower()
    if any(keyword in user_message_lower for keyword in FUNCTION_KEYWORDS):
        await help_command(update, context)
        return

    # Nếu chỉ tag bot
    if bot_username in user_message:
        if len(user_message.replace(bot_username, "").strip()) == 0:
            reply = handle_tag_only()
        else:
            reply = await get_grok_response(user_message.replace(bot_username, "").strip(), chat_id, user_id, is_group_context=is_group_chat)
        await context.bot.send_message(chat_id=chat_id, text=reply)
        return

    # Nếu trả lời lại tin nhắn của bot
    if update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id:
        user_reply = user_message
        is_group_chat = update.message.chat.type in ["group", "supergroup"]  # THÊM DÒNG NÀY
        reply = await get_grok_response(user_reply, chat_id, user_id, is_group_context=is_group_chat)
        await context.bot.send_message(chat_id=chat_id, text=reply)
        return

# Hàm gửi tin nhắn tự động
async def send_daily_message(application):
    for chat_id in ALLOWED_CHAT_IDS:
        try:
            await application.bot.send_message(chat_id=chat_id, text="Hí ae, trưa rồi, chúc ae ăn ngon, nghỉ khỏe. Buổi trưa vui vẻ nha ae!😎")
            print(f"Đã gửi tin nhắn cho chat {chat_id}")
        except Exception as e:
            print(f"Lỗi khi gửi tin nhắn: {e}")

# Hàm khởi tạo scheduler
def start_scheduler(application):
    scheduler = AsyncIOScheduler()
    tz = pytz.timezone('Asia/Ho_Chi_Minh')
    scheduler.add_job(send_daily_message, 'cron', hour=12, minute=00, args=[application], timezone=tz)
    scheduler.add_job(cleanup_group_history, 'interval', minutes=10)
    scheduler.start()
    print("Scheduler đã được khởi tạo và đang chạy...")

# Hàm main với logic lọc tin nhắn và tối ưu độ trễ
async def main():
    application = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    application.add_handler(MessageHandler(filters.ALL & (~filters.COMMAND), handle_message))
    application.add_handler(CommandHandler("search", search_command))
    application.add_handler(CommandHandler("e", e_command))
    application.add_handler(CommandHandler("checklimit", check_limit_command))
    application.add_handler(CommandHandler("dh", dh_command))
    add_help_handler(application)

    # Khởi tạo Application
    await application.initialize()
    print("Application đã được khởi tạo!")

    # Khởi tạo scheduler
    start_scheduler(application)
    print("Bot đã khởi động và lịch trình đã được thiết lập!")

    try:
        # Lấy tối đa 100 updates từ Telegram để giảm độ trễ
        updates = []
        offset = None
        max_updates = 100
        while len(updates) < max_updates:
            batch = await application.bot.get_updates(offset=offset, timeout=1)
            if not batch:
                break
            updates.extend(batch)
            offset = max(update.update_id for update in batch) + 1

        # Lọc các updates chỉ chứa tin nhắn
        message_updates = []
        for u in updates:
            if u.message and hasattr(u.message, 'date'):
                message_updates.append(u)
            else:
                print(f"Bỏ qua update {u.update_id}: Không có message hoặc date")

        # Lấy thời gian hiện tại (múi giờ UTC)
        current_time = datetime.now(pytz.UTC)
        time_threshold = current_time - timedelta(minutes=5)

        # Lọc tin nhắn trong 5 phút gần nhất
        recent_updates = []
        for u in message_updates:
            try:
                if isinstance(u.message.date, datetime):
                    message_time = u.message.date.replace(tzinfo=pytz.UTC)
                else:
                    message_time = datetime.fromtimestamp(u.message.date, tz=pytz.UTC)

                if message_time >= time_threshold:
                    recent_updates.append(u)
            except Exception as e:
                continue

        # Nếu không có tin nhắn trong 5 phút, lấy tối đa 5 tin nhắn gần nhất
        if not recent_updates:
            message_updates.sort(key=lambda x: x.update_id)
            recent_updates = message_updates[-5:] if len(message_updates) > 5 else message_updates

        print(f"Tổng số tin nhắn trong hàng đợi: {len(message_updates)}")

        # Xử lý từng tin nhắn đã lọc
        for update in recent_updates:
            try:
                if isinstance(update.message.date, datetime):
                    message_time = update.message.date.replace(tzinfo=pytz.UTC)
                else:
                    message_time = datetime.fromtimestamp(update.message.date, tz=pytz.UTC)
                context = ContextTypes.DEFAULT_TYPE(application=application)
                await handle_message(update, context)
            except Exception as e:
                pass

        # Đánh dấu toàn bộ hàng đợi là đã xử lý
        if updates:
            last_update_id = max(update.update_id for update in updates)
            await application.bot.get_updates(offset=last_update_id + 1, timeout=1)

    except Exception as e:
        pass

    # Bắt đầu polling để xử lý tin nhắn mới
    try:
        await application.run_polling(allowed_updates=["message"])
    finally:
        await application.shutdown()
        print("Application đã được tắt!")

# Khởi động bot
nest_asyncio.apply()
asyncio.run(main())
if __name__ == "__main__":
    nest_asyncio.apply()
    asyncio.run(main())
